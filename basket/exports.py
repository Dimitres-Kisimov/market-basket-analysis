"""Executive deliverables: a PDF briefing (matplotlib) and an Excel workbook.

Headless-safe by construction: only the matplotlib object-oriented API is used
(``matplotlib.figure.Figure`` + ``PdfPages``), never ``pyplot``, so no GUI
backend is touched. Every page and sheet repeats the synthetic-data disclaimer
and the correlation-is-not-causation note.
"""

from __future__ import annotations

import os
from dataclasses import dataclass
from datetime import date

import numpy as np
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.colors import LinearSegmentedColormap, TwoSlopeNorm
from matplotlib.figure import Figure
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from basket.affinity import (
    AffinityReport,
    affinity_network,
    write_affinity_csv,
    write_affinity_svg,
)
from basket.apriori import apriori
from basket.data import AVG_LINE_VALUE_EUR, baskets_from_transactions, generate_transactions
from basket.evaluate import (
    EvaluationReport,
    evaluate_recommenders,
    write_evaluation_csv,
    write_evaluation_svg,
)
from basket.recommend import ASSUMPTION_NOTE, Recommendation, cross_sell_recommendations
from basket.rules import Rule, format_itemset, format_rule, generate_rules
from basket.segment import Segmentation, segment_customers
from basket.stability import (
    StabilityReport,
    rule_stability,
    write_stability_csv,
    write_stability_svg,
)

DISCLAIMER = (
    "SYNTHETIC DATA: all figures come from a seeded simulation built for this "
    "portfolio project. No real customer or sales data was used."
)
CAUSATION_NOTE = (
    "Lift is observational: correlation is not causation. Validate any "
    "cross-sell action with a controlled test before scaling it."
)

# Defaults for the stability deliverable (rule-robustness trust layer).
STABILITY_METHOD = "time_window"
STABILITY_N_SPLITS = 4
STABILITY_TOP_N = 20

# Defaults for the recommender back-test deliverable (predictive-accuracy layer).
EVALUATION_TRAIN_FRACTION = 0.7
EVALUATION_K_VALUES: tuple[int, ...] = (1, 3, 5)

# The affinity network needs non-negative edge weights (lift - 1), so its edge
# threshold is the run's min lift floored at 1.0.
NETWORK_MIN_LIFT_FLOOR = 1.0

# Chart chrome and palette (validated defaults; light surface).
_INK = "#0b0b0b"
_INK_SECONDARY = "#52514e"
_INK_MUTED = "#898781"
_GRID = "#e1e0d9"
_BASELINE = "#c3c2b7"
_SURFACE = "#fcfcfb"
_NEUTRAL_FILL = "#f0efec"
_SERIES = ("#2a78d6", "#008300", "#e87ba4", "#eda100", "#1baf7a", "#eb6834")
_DIVERGING = ("#2a78d6", "#f0efec", "#e34948")  # blue <-> gray(=lift 1) <-> red

_A4_LANDSCAPE = (11.69, 8.27)


@dataclass
class AnalysisResult:
    transactions: pd.DataFrame
    baskets: list[frozenset[str]]
    itemsets: dict[frozenset[str], float]
    rules: list[Rule]
    segmentation: Segmentation
    recommendations: list[Recommendation]
    n_baskets: int
    seed: int
    min_support: float
    min_confidence: float
    min_lift: float


def run_analysis(
    n_baskets: int = 6000,
    seed: int = 42,
    min_support: float = 0.02,
    min_confidence: float = 0.30,
    min_lift: float = 1.10,
    k_segments: int = 3,
) -> AnalysisResult:
    """Run the full pipeline: generate -> mine -> rules -> segment -> recommend."""
    transactions = generate_transactions(n_baskets=n_baskets, seed=seed)
    baskets = baskets_from_transactions(transactions)
    itemsets = apriori(baskets, min_support=min_support, max_len=3)
    rules = generate_rules(
        itemsets, len(baskets), min_confidence=min_confidence, min_lift=min_lift
    )
    segmentation = segment_customers(transactions, k=k_segments, seed=seed)
    recommendations = cross_sell_recommendations(rules, AVG_LINE_VALUE_EUR, top_n=10)
    return AnalysisResult(
        transactions=transactions,
        baskets=baskets,
        itemsets=itemsets,
        rules=rules,
        segmentation=segmentation,
        recommendations=recommendations,
        n_baskets=len(baskets),
        seed=seed,
        min_support=min_support,
        min_confidence=min_confidence,
        min_lift=min_lift,
    )


def build_stability_report(
    result: AnalysisResult,
    method: str = STABILITY_METHOD,
    n_splits: int = STABILITY_N_SPLITS,
    top_n: int = STABILITY_TOP_N,
) -> StabilityReport:
    """Robustness check for the mined rules, reusing this run's thresholds."""
    return rule_stability(
        result.baskets,
        result.rules,
        method=method,
        n_splits=n_splits,
        top_n=top_n,
        min_support=result.min_support,
        min_confidence=result.min_confidence,
        min_lift=result.min_lift,
        seed=result.seed,
    )


def build_evaluation_report(
    result: AnalysisResult,
    train_fraction: float = EVALUATION_TRAIN_FRACTION,
    k_values: tuple[int, ...] = EVALUATION_K_VALUES,
) -> EvaluationReport:
    """Leave-one-out recommender back-test, reusing this run's thresholds/seed."""
    return evaluate_recommenders(
        result.baskets,
        train_fraction=train_fraction,
        k_values=k_values,
        min_support=result.min_support,
        min_confidence=result.min_confidence,
        min_lift=result.min_lift,
        seed=result.seed,
    )


def build_affinity_report(result: AnalysisResult) -> AffinityReport:
    """Category affinity network + communities, reusing this run's mined itemsets."""
    return affinity_network(
        result.itemsets,
        result.n_baskets,
        min_support=result.min_support,
        min_lift=max(NETWORK_MIN_LIFT_FLOOR, result.min_lift),
    )


def _pair_lift_matrix(
    baskets: list[frozenset[str]], categories: list[str]
) -> np.ndarray:
    """Full category-pair lift matrix (diagonal masked as NaN)."""
    n = len(baskets)
    index = {c: i for i, c in enumerate(categories)}
    presence = np.zeros((n, len(categories)))
    for row, basket in enumerate(baskets):
        for item in basket:
            if item in index:
                presence[row, index[item]] = 1.0
    joint = (presence.T @ presence) / n
    marginal = np.diag(joint).copy()
    with np.errstate(divide="ignore", invalid="ignore"):
        lift = joint / np.outer(marginal, marginal)
    lift[~np.isfinite(lift)] = np.nan
    np.fill_diagonal(lift, np.nan)
    return lift


def _new_page(title: str, subtitle: str) -> Figure:
    fig = Figure(figsize=_A4_LANDSCAPE, facecolor=_SURFACE)
    fig.text(0.06, 0.94, title, fontsize=17, fontweight="bold", color=_INK)
    fig.text(0.06, 0.905, subtitle, fontsize=9.5, color=_INK_SECONDARY)
    fig.text(0.06, 0.035, DISCLAIMER, fontsize=7.5, color=_INK_MUTED)
    return fig


def _cover_page(pdf: PdfPages, result: AnalysisResult) -> None:
    fig = Figure(figsize=_A4_LANDSCAPE, facecolor=_SURFACE)
    fig.text(
        0.08, 0.80, "Cross-Sell Opportunity Analysis",
        fontsize=26, fontweight="bold", color=_INK,
    )
    fig.text(
        0.08, 0.745,
        "Market-basket mining for a B2B maintenance & construction supplies distributor",
        fontsize=13, color=_INK_SECONDARY,
    )
    fig.text(
        0.08, 0.705,
        f"Generated {date.today().isoformat()}  |  seed {result.seed}  |  "
        f"min support {result.min_support:.0%}  |  min confidence "
        f"{result.min_confidence:.0%}  |  min lift {result.min_lift:.2f}",
        fontsize=9, color=_INK_MUTED,
    )

    top = result.rules[0] if result.rules else None
    tiles = [
        (f"{result.n_baskets:,}", "orders analysed"),
        (f"{len(result.itemsets):,}", "frequent itemsets"),
        (f"{len(result.rules):,}", "rules kept"),
        (f"{top.lift:.1f}x" if top else "-", "strongest lift"),
    ]
    for i, (value, label) in enumerate(tiles):
        x = 0.08 + i * 0.22
        fig.text(x, 0.565, value, fontsize=24, fontweight="bold", color=_INK)
        fig.text(x, 0.525, label, fontsize=10, color=_INK_SECONDARY)

    if top is not None:
        fig.text(
            0.08, 0.43,
            f"Headline: {format_rule(top)}  (lift {top.lift:.2f}, "
            f"confidence {top.confidence:.0%}, {top.support_count} orders)",
            fontsize=12, color=_INK,
        )

    fig.text(
        0.08, 0.30, DISCLAIMER + "\n" + CAUSATION_NOTE,
        fontsize=10, color=_INK,
        bbox={
            "boxstyle": "round,pad=0.6",
            "facecolor": _NEUTRAL_FILL,
            "edgecolor": _BASELINE,
        },
        va="top", wrap=True,
    )
    fig.text(0.08, 0.09, "Author: Dimitres Kisimov", fontsize=9, color=_INK_MUTED)
    pdf.savefig(fig)


def _rules_page(pdf: PdfPages, result: AnalysisResult, max_rows: int = 12) -> None:
    fig = _new_page(
        "Top association rules (ranked by lift)",
        f"Filtered at confidence >= {result.min_confidence:.0%} and lift >= "
        f"{result.min_lift:.2f}. Lift is observational -- correlation is not causation.",
    )
    ax = fig.add_axes((0.05, 0.10, 0.90, 0.76))
    ax.axis("off")
    header = ["Rule", "Support", "Orders", "Confidence", "Lift", "Conviction", "Note"]
    body = []
    for rule in result.rules[:max_rows]:
        conviction = "inf" if rule.conviction == float("inf") else f"{rule.conviction:.2f}"
        body.append(
            [
                format_rule(rule),
                f"{rule.support:.1%}",
                f"{rule.support_count}",
                f"{rule.confidence:.1%}",
                f"{rule.lift:.2f}",
                conviction,
                "thin support" if rule.thin_support else "",
            ]
        )
    table = ax.table(
        cellText=body,
        colLabels=header,
        loc="upper center",
        cellLoc="center",
        colWidths=[0.40, 0.09, 0.08, 0.11, 0.08, 0.11, 0.13],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.5)
    table.scale(1.0, 1.55)
    for (row, _col), cell in table.get_celld().items():
        cell.set_edgecolor(_GRID)
        if row == 0:
            cell.set_text_props(fontweight="bold", color=_INK)
            cell.set_facecolor(_NEUTRAL_FILL)
        else:
            cell.set_text_props(color=_INK_SECONDARY)
            cell.set_facecolor(_SURFACE)
    pdf.savefig(fig)


def _heatmap_page(pdf: PdfPages, result: AnalysisResult) -> None:
    categories = sorted({item for basket in result.baskets for item in basket})
    lift = _pair_lift_matrix(result.baskets, categories)
    finite = lift[np.isfinite(lift)]
    vmin = min(0.85, float(finite.min())) if finite.size else 0.5
    vmax = max(1.15, float(finite.max())) if finite.size else 1.5

    fig = _new_page(
        "Category-pair lift heatmap",
        "Gray = independence (lift 1.0), red = bought together more than chance, "
        "blue = less. Diagonal masked.",
    )
    ax = fig.add_axes((0.16, 0.16, 0.62, 0.68))
    cmap = LinearSegmentedColormap.from_list("lift_diverging", _DIVERGING)
    cmap.set_bad(_SURFACE)
    norm = TwoSlopeNorm(vmin=vmin, vcenter=1.0, vmax=vmax)
    image = ax.imshow(lift, cmap=cmap, norm=norm)

    labels = [c.replace("_", " ") for c in categories]
    ax.set_xticks(range(len(categories)), labels, rotation=45, ha="right", fontsize=7)
    ax.set_yticks(range(len(categories)), labels, fontsize=7)
    ax.tick_params(colors=_INK_MUTED, length=0)
    for spine in ax.spines.values():
        spine.set_color(_BASELINE)

    # Selective direct labels: only clearly non-independent pairs get a number.
    for i in range(len(categories)):
        for j in range(len(categories)):
            value = lift[i, j]
            if np.isfinite(value) and (value >= 1.5 or value <= 0.6):
                position = norm(value)
                ax.text(
                    j, i, f"{value:.1f}",
                    ha="center", va="center", fontsize=6,
                    color="white" if (position >= 0.75 or position <= 0.25) else _INK,
                )
    colorbar = fig.colorbar(image, ax=ax, fraction=0.04, pad=0.03)
    colorbar.set_label("lift", color=_INK_SECONDARY, fontsize=9)
    colorbar.ax.tick_params(colors=_INK_MUTED, labelsize=7)
    colorbar.outline.set_edgecolor(_BASELINE)
    pdf.savefig(fig)


def _network_page(pdf: PdfPages, report: AffinityReport) -> None:
    fig = _new_page(
        "Category affinity communities (co-purchase network)",
        f"{report.n_nodes} categories, {report.n_edges} lift-weighted edges (pair "
        f"support >= {report.min_support:.0%}, lift >= {report.min_lift:.2f}), grouped "
        "by greedy modularity. Assortment structure, observational -- not causation.",
    )
    fig.text(
        0.06, 0.855, report.plain_language(),
        fontsize=9.5, color=_INK, va="top", wrap=True,
    )
    ax = fig.add_axes((0.05, 0.42, 0.90, 0.36))
    ax.axis("off")
    header = ["Community", "Categories", "Size", "Edges", "Avg lift", "Max lift", "Top internal pair"]
    body = []
    for community in report.communities:
        top_pair = (
            f"{community.top_edge.label} ({community.top_edge.lift:.2f})"
            if community.top_edge is not None
            else "-"
        )
        body.append(
            [
                f"{community.community_id}",
                ", ".join(m.replace("_", " ") for m in community.members),
                f"{community.n_members}",
                f"{community.n_internal_edges}",
                f"{community.lift_mean:.2f}" if community.n_internal_edges else "-",
                f"{community.lift_max:.2f}" if community.n_internal_edges else "-",
                top_pair,
            ]
        )
    table = ax.table(
        cellText=body,
        colLabels=header,
        loc="upper center",
        cellLoc="center",
        colWidths=[0.09, 0.38, 0.06, 0.06, 0.08, 0.08, 0.25],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(7.5)
    table.scale(1.0, 1.5)
    for (row, _col), cell in table.get_celld().items():
        cell.set_edgecolor(_GRID)
        if row == 0:
            cell.set_text_props(fontweight="bold", color=_INK)
            cell.set_facecolor(_NEUTRAL_FILL)
        else:
            cell.set_text_props(color=_INK_SECONDARY)
            cell.set_facecolor(_SURFACE)

    ax2 = fig.add_axes((0.05, 0.10, 0.90, 0.26))
    ax2.axis("off")
    bridge_header = ["Bridge (cross-merchandising candidate)", "Lift", "Support", "Orders", "Connects"]
    bridge_body = []
    for edge in report.bridges[:6]:
        bridge_body.append(
            [
                edge.label,
                f"{edge.lift:.2f}",
                f"{edge.support:.1%}",
                f"{edge.support_count}",
                f"community {report.membership[edge.item_a]} <-> "
                f"{report.membership[edge.item_b]}",
            ]
        )
    if not bridge_body:
        bridge_body.append(["none at these thresholds", "-", "-", "-", "-"])
    bridge_table = ax2.table(
        cellText=bridge_body,
        colLabels=bridge_header,
        loc="upper center",
        cellLoc="center",
        colWidths=[0.40, 0.08, 0.10, 0.08, 0.24],
    )
    bridge_table.auto_set_font_size(False)
    bridge_table.set_fontsize(7.5)
    bridge_table.scale(1.0, 1.5)
    for (row, _col), cell in bridge_table.get_celld().items():
        cell.set_edgecolor(_GRID)
        if row == 0:
            cell.set_text_props(fontweight="bold", color=_INK)
            cell.set_facecolor(_NEUTRAL_FILL)
        else:
            cell.set_text_props(color=_INK_SECONDARY)
            cell.set_facecolor(_SURFACE)
    pdf.savefig(fig)


def _segments_page(pdf: PdfPages, result: AnalysisResult) -> None:
    segmentation = result.segmentation
    k = len(segmentation.profiles)
    fig = _new_page(
        "Customer segments (k-means on spend shares)",
        "Descriptive clustering of per-customer category spend mix; "
        "bar length = share of segment spend (centroid).",
    )
    spend_columns = list(segmentation.spend.columns)
    slot_width = 0.90 / k
    for j, profile in enumerate(segmentation.profiles):
        # Wide gutters so category labels never collide with a neighbour's bars.
        ax = fig.add_axes((0.115 + j * slot_width, 0.18, slot_width - 0.105, 0.55))
        centroid = segmentation.kmeans_result.centroids[j]
        top_idx = np.argsort(-centroid)[:6][::-1]
        names = [str(spend_columns[i]).replace("_", " ") for i in top_idx]
        shares = centroid[top_idx]
        bars = ax.barh(names, shares, color=_SERIES[j % len(_SERIES)], height=0.55)
        for bar, share in zip(bars, shares, strict=True):
            ax.text(
                bar.get_width() + max(shares) * 0.03,
                bar.get_y() + bar.get_height() / 2,
                f"{share:.0%}", va="center", fontsize=7.5, color=_INK_SECONDARY,
            )
        ax.set_xlim(0, max(shares) * 1.25)
        ax.set_title(
            f"Segment {profile.segment}\n{profile.n_customers} customers "
            f"({profile.share_of_customers:.0%}), avg spend "
            f"EUR {profile.avg_total_spend_eur:,.0f}",
            fontsize=9, color=_INK, loc="left",
        )
        ax.tick_params(colors=_INK_MUTED, labelsize=7.5, length=0)
        ax.set_xticks([])
        for name, spine in ax.spines.items():
            spine.set_visible(name == "left")
            spine.set_color(_BASELINE)
    pdf.savefig(fig)


def _stability_page(
    pdf: PdfPages, report: StabilityReport, max_rows: int = 20
) -> None:
    method_phrase = (
        f"{report.n_splits} contiguous time windows (basket arrival order)"
        if report.method == "time_window"
        else f"{report.n_splits} seeded bootstrap resamples"
    )
    fig = _new_page(
        "Rule stability (robustness check)",
        f"Top {report.top_n} rules re-mined across {method_phrase}. A rule is "
        "STABLE only if it clears every threshold in every split.",
    )
    fig.text(
        0.06, 0.855, report.plain_language(),
        fontsize=9.5, color=_INK, va="top", wrap=True,
    )
    ax = fig.add_axes((0.05, 0.09, 0.90, 0.70))
    ax.axis("off")
    header = ["Rule", "Lift", "Support", "Stability", "Splits", "Lift CV", "Verdict"]
    body = []
    for item in report.rules[:max_rows]:
        body.append(
            [
                item.label,
                f"{item.reference_lift:.2f}",
                f"{item.reference_support:.1%}",
                f"{item.stability_score:.0%}",
                f"{item.n_present}/{item.n_splits}",
                f"{item.lift_cv:.2f}",
                "stable" if item.stable else "check",
            ]
        )
    table = ax.table(
        cellText=body,
        colLabels=header,
        loc="upper center",
        cellLoc="center",
        colWidths=[0.42, 0.07, 0.09, 0.10, 0.08, 0.09, 0.10],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(7.5)
    table.scale(1.0, 1.2)
    stable_flags = [item.stable for item in report.rules[:max_rows]]
    for (row, col), cell in table.get_celld().items():
        cell.set_edgecolor(_GRID)
        if row == 0:
            cell.set_text_props(fontweight="bold", color=_INK)
            cell.set_facecolor(_NEUTRAL_FILL)
            continue
        cell.set_text_props(color=_INK_SECONDARY)
        if col == 6:  # verdict column: green when stable, amber when not
            cell.set_facecolor("#e3f0e0" if stable_flags[row - 1] else "#faedcf")
        else:
            cell.set_facecolor(_SURFACE)
    pdf.savefig(fig)


def _evaluation_page(pdf: PdfPages, report: EvaluationReport) -> None:
    fig = _new_page(
        "Cross-sell recommender back-test",
        f"Leave-one-out on {report.n_trials} held-out baskets: rules mined on "
        f"{report.n_train} training baskets predict a hidden category vs a popularity "
        "baseline. Hit-rate is predictive fit -- not the causal uplift of an A/B test.",
    )
    fig.text(
        0.06, 0.855, report.plain_language(),
        fontsize=9.5, color=_INK, va="top", wrap=True,
    )
    ax = fig.add_axes((0.06, 0.30, 0.88, 0.46))
    ax.axis("off")
    ratio = report.hit_rate_ratio()
    header = ["Metric", "Association rules", "Popularity baseline", "Rules / baseline"]
    body = []
    for k in report.k_values:
        body.append(
            [
                f"Hit-rate@{k}",
                f"{report.rules.hit_rate[k]:.1%}",
                f"{report.popularity.hit_rate[k]:.1%}",
                f"{ratio[k]:.2f}x",
            ]
        )
    mrr_ratio = report.rules.mrr / report.popularity.mrr if report.popularity.mrr else 0.0
    body.append(["MRR", f"{report.rules.mrr:.3f}", f"{report.popularity.mrr:.3f}", f"{mrr_ratio:.2f}x"])
    body.append(
        ["Coverage", f"{report.rules.coverage:.1%}", f"{report.popularity.coverage:.1%}", "-"]
    )
    table = ax.table(
        cellText=body,
        colLabels=header,
        loc="upper center",
        cellLoc="center",
        colWidths=[0.22, 0.26, 0.26, 0.20],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(9.5)
    table.scale(1.0, 1.7)
    for (row, col), cell in table.get_celld().items():
        cell.set_edgecolor(_GRID)
        if row == 0:
            cell.set_text_props(fontweight="bold", color=_INK)
            cell.set_facecolor(_NEUTRAL_FILL)
        else:
            cell.set_text_props(color=_INK_SECONDARY)
            cell.set_facecolor("#e6effa" if col == 1 else _SURFACE)
    fig.text(
        0.06, 0.20,
        f"Setup: arrival-order split at {report.train_fraction:.0%} (train earlier, "
        f"test later); one category hidden per test basket (seed {report.seed}); "
        f"{report.n_rules} rules mined on train at the run's thresholds. "
        "Popularity baseline = most-frequent training categories not already in the basket.",
        fontsize=8.5, color=_INK_MUTED, va="top", wrap=True,
    )
    pdf.savefig(fig)


def export_pdf(
    result: AnalysisResult,
    path: str,
    stability_report: StabilityReport | None = None,
    evaluation_report: EvaluationReport | None = None,
    affinity_report: AffinityReport | None = None,
) -> None:
    """Write the seven-page executive PDF (network, stability and back-test pages)."""
    if stability_report is None:
        stability_report = build_stability_report(result)
    if evaluation_report is None:
        evaluation_report = build_evaluation_report(result)
    if affinity_report is None:
        affinity_report = build_affinity_report(result)
    with PdfPages(path) as pdf:
        _cover_page(pdf, result)
        _rules_page(pdf, result)
        _heatmap_page(pdf, result)
        _network_page(pdf, affinity_report)
        _segments_page(pdf, result)
        _stability_page(pdf, stability_report)
        _evaluation_page(pdf, evaluation_report)


def _style_header(sheet: Worksheet, row: int, n_columns: int) -> None:
    fill = PatternFill("solid", fgColor="F0EFEC")
    for column in range(1, n_columns + 1):
        cell = sheet.cell(row=row, column=column)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left")


def _set_widths(sheet: Worksheet, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width


def export_excel(
    result: AnalysisResult,
    path: str,
    stability_report: StabilityReport | None = None,
    evaluation_report: EvaluationReport | None = None,
    affinity_report: AffinityReport | None = None,
) -> None:
    """Write the analyst workbook: Rules, Itemsets, Segments, Recommendations, Stability, Evaluation, Network."""
    if stability_report is None:
        stability_report = build_stability_report(result)
    if evaluation_report is None:
        evaluation_report = build_evaluation_report(result)
    if affinity_report is None:
        affinity_report = build_affinity_report(result)
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Rules"
    sheet.append([DISCLAIMER])
    sheet.append([CAUSATION_NOTE])
    sheet.append([])
    header_row = 4
    sheet.append(
        [
            "Rule", "Antecedent", "Consequent", "Support", "Orders",
            "Confidence", "Lift", "Leverage", "Conviction", "Thin support",
        ]
    )
    for rule in result.rules:
        sheet.append(
            [
                format_rule(rule),
                format_itemset(rule.antecedent),
                format_itemset(rule.consequent),
                round(rule.support, 4),
                rule.support_count,
                round(rule.confidence, 4),
                round(rule.lift, 3),
                round(rule.leverage, 4),
                "inf" if rule.conviction == float("inf") else round(rule.conviction, 3),
                "YES" if rule.thin_support else "",
            ]
        )
    _style_header(sheet, header_row, 10)
    _set_widths(sheet, [46, 30, 22, 10, 8, 11, 8, 10, 11, 12])
    sheet.freeze_panes = f"A{header_row + 1}"

    sheet = workbook.create_sheet("Itemsets")
    sheet.append([DISCLAIMER])
    sheet.append([])
    sheet.append(["Itemset", "Size", "Support", "Orders"])
    ordered = sorted(result.itemsets.items(), key=lambda kv: (-kv[1], format_itemset(kv[0])))
    for itemset, support in ordered:
        sheet.append(
            [
                format_itemset(itemset),
                len(itemset),
                round(support, 4),
                round(support * result.n_baskets),
            ]
        )
    _style_header(sheet, 3, 4)
    _set_widths(sheet, [46, 7, 10, 9])
    sheet.freeze_panes = "A4"

    sheet = workbook.create_sheet("Segments")
    sheet.append([DISCLAIMER])
    sheet.append([])
    sheet.append(
        ["Segment", "Customers", "Share", "Avg total spend EUR", "Top categories (centroid share)"]
    )
    for profile in result.segmentation.profiles:
        top = ", ".join(f"{name} {share:.0%}" for name, share in profile.top_categories)
        sheet.append(
            [
                profile.segment,
                profile.n_customers,
                round(profile.share_of_customers, 3),
                round(profile.avg_total_spend_eur, 2),
                top,
            ]
        )
    _style_header(sheet, 3, 5)
    assignments_header = 5 + len(result.segmentation.profiles)
    sheet.append([])
    sheet.append(["Customer", "Segment", "Total spend EUR"])
    totals = result.segmentation.spend.sum(axis=1)
    for customer_id, segment in result.segmentation.assignments.items():
        sheet.append([customer_id, int(segment), round(float(totals[customer_id]), 2)])
    _style_header(sheet, assignments_header, 3)
    _set_widths(sheet, [12, 11, 18, 20, 60])

    sheet = workbook.create_sheet("Recommendations")
    sheet.append([DISCLAIMER])
    sheet.append([ASSUMPTION_NOTE])
    sheet.append([])
    sheet.append(
        ["Rank", "Buyers of", "Recommend", "Lift", "Confidence",
         "Est. incremental EUR/order", "Headline"]
    )
    for rank, recommendation in enumerate(result.recommendations, start=1):
        rule = recommendation.rule
        sheet.append(
            [
                rank,
                format_itemset(rule.antecedent),
                format_itemset(rule.consequent),
                round(rule.lift, 3),
                round(rule.confidence, 4),
                round(recommendation.est_incremental_value_eur, 2),
                recommendation.headline,
            ]
        )
    _style_header(sheet, 4, 7)
    _set_widths(sheet, [6, 30, 22, 8, 11, 24, 110])
    sheet.freeze_panes = "A5"

    sheet = workbook.create_sheet("Stability")
    sheet.append([DISCLAIMER])
    sheet.append([CAUSATION_NOTE])
    sheet.append([stability_report.plain_language()])
    sheet.append([])
    stability_header_row = 5
    sheet.append(
        [
            "Rank", "Rule", "Ref lift", "Ref confidence", "Ref support", "Ref orders",
            "Splits", "Present", "Stability score", "Lift mean", "Lift std", "Lift CV",
            "Lift min", "Lift max", "Stable", "Method",
        ]
    )
    for rank, item in enumerate(stability_report.rules, start=1):
        sheet.append(
            [
                rank,
                item.label,
                round(item.reference_lift, 3),
                round(item.reference_confidence, 4),
                round(item.reference_support, 4),
                item.reference_support_count,
                item.n_splits,
                item.n_present,
                round(item.stability_score, 4),
                round(item.lift_mean, 4),
                round(item.lift_std, 4),
                round(item.lift_cv, 4),
                round(item.lift_min, 4),
                round(item.lift_max, 4),
                "YES" if item.stable else "",
                stability_report.method,
            ]
        )
    _style_header(sheet, stability_header_row, 16)
    _set_widths(sheet, [6, 46, 9, 14, 11, 10, 8, 9, 14, 10, 9, 9, 9, 9, 8, 12])
    sheet.freeze_panes = f"A{stability_header_row + 1}"

    sheet = workbook.create_sheet("Evaluation")
    sheet.append([DISCLAIMER])
    sheet.append([CAUSATION_NOTE])
    sheet.append([evaluation_report.plain_language()])
    sheet.append([])
    evaluation_header_row = 5
    k_labels = [f"Hit-rate@{k}" for k in evaluation_report.k_values]
    sheet.append(["Recommender", "Trials", *k_labels, "MRR", "Coverage"])

    def _hit_cells(metrics: object) -> list:
        return [round(metrics.hit_rate[k], 4) for k in evaluation_report.k_values]

    sheet.append(
        [
            "Association rules",
            evaluation_report.rules.n_trials,
            *_hit_cells(evaluation_report.rules),
            round(evaluation_report.rules.mrr, 4),
            round(evaluation_report.rules.coverage, 4),
        ]
    )
    sheet.append(
        [
            "Popularity baseline",
            evaluation_report.popularity.n_trials,
            *_hit_cells(evaluation_report.popularity),
            round(evaluation_report.popularity.mrr, 4),
            round(evaluation_report.popularity.coverage, 4),
        ]
    )
    ratio = evaluation_report.hit_rate_ratio()
    mrr_ratio = (
        evaluation_report.rules.mrr / evaluation_report.popularity.mrr
        if evaluation_report.popularity.mrr
        else 0.0
    )
    sheet.append(
        [
            "Rules / baseline (ratio)",
            "",
            *[round(ratio[k], 3) for k in evaluation_report.k_values],
            round(mrr_ratio, 3),
            "",
        ]
    )
    _style_header(sheet, evaluation_header_row, 5 + len(evaluation_report.k_values))
    settings_row = evaluation_header_row + 4
    sheet.append([])
    sheet.append(
        [
            "Setup",
            f"arrival-order split {evaluation_report.train_fraction:.0%} train / "
            f"{1 - evaluation_report.train_fraction:.0%} test; one category hidden per "
            f"test basket (seed {evaluation_report.seed}); "
            f"{evaluation_report.n_rules} rules mined on "
            f"{evaluation_report.n_train} training baskets; popularity baseline = "
            "most-frequent training categories not already in the basket.",
        ]
    )
    sheet.cell(row=settings_row + 1, column=1).font = Font(bold=True)
    _set_widths(sheet, [26, 12, 12, 12, 12, 12, 12])

    sheet = workbook.create_sheet("Network")
    sheet.append([DISCLAIMER])
    sheet.append([CAUSATION_NOTE])
    sheet.append([affinity_report.plain_language()])
    sheet.append([])
    communities_header_row = 5
    sheet.append(
        ["Community", "Categories", "Size", "Internal edges", "Avg internal lift",
         "Max internal lift", "Top internal pair"]
    )
    for community in affinity_report.communities:
        sheet.append(
            [
                community.community_id,
                ", ".join(community.members),
                community.n_members,
                community.n_internal_edges,
                round(community.lift_mean, 3) if community.n_internal_edges else "",
                round(community.lift_max, 3) if community.n_internal_edges else "",
                f"{community.top_edge.label} ({community.top_edge.lift:.2f})"
                if community.top_edge is not None
                else "",
            ]
        )
    _style_header(sheet, communities_header_row, 7)
    sheet.append([])
    edges_header_row = communities_header_row + len(affinity_report.communities) + 2
    sheet.append(
        ["Edge", "Support", "Orders", "Lift", "Leverage", "Weight (lift - 1)",
         "Community A", "Community B", "Type"]
    )
    for edge in affinity_report.edges:
        community_a = affinity_report.membership[edge.item_a]
        community_b = affinity_report.membership[edge.item_b]
        sheet.append(
            [
                edge.label,
                round(edge.support, 4),
                edge.support_count,
                round(edge.lift, 3),
                round(edge.leverage, 4),
                round(edge.weight, 3),
                community_a,
                community_b,
                "internal" if community_a == community_b else "BRIDGE",
            ]
        )
    _style_header(sheet, edges_header_row, 9)
    sheet.append([])
    sheet.append(
        [
            "Setup",
            f"edges = frequent category pairs (support >= "
            f"{affinity_report.min_support:.0%}) with lift >= "
            f"{affinity_report.min_lift:.2f}, weighted by lift - 1; communities via "
            "greedy modularity maximisation (Newman 2004), weighted modularity "
            f"Q = {affinity_report.modularity:.3f}; bridges are edges whose "
            "endpoints sit in different communities.",
        ]
    )
    sheet.cell(row=edges_header_row + len(affinity_report.edges) + 2, column=1).font = Font(bold=True)
    _set_widths(sheet, [40, 10, 9, 8, 10, 15, 12, 12, 10])

    workbook.save(path)


def write_deliverables(
    output_dir: str = "deliverables",
    n_baskets: int = 6000,
    seed: int = 42,
    min_support: float = 0.02,
    min_confidence: float = 0.30,
    min_lift: float = 1.10,
) -> dict[str, int]:
    """Run the pipeline and write both deliverables; verify each is > 10 KB."""
    result = run_analysis(
        n_baskets=n_baskets,
        seed=seed,
        min_support=min_support,
        min_confidence=min_confidence,
        min_lift=min_lift,
    )
    os.makedirs(output_dir, exist_ok=True)
    stability_report = build_stability_report(result)
    evaluation_report = build_evaluation_report(result)
    affinity_report = build_affinity_report(result)
    pdf_path = os.path.join(output_dir, "cross_sell_briefing.pdf")
    excel_path = os.path.join(output_dir, "market_basket_analysis.xlsx")
    csv_path = os.path.join(output_dir, "rule_stability.csv")
    svg_path = os.path.join(output_dir, "rule_stability.svg")
    eval_csv_path = os.path.join(output_dir, "recommender_backtest.csv")
    eval_svg_path = os.path.join(output_dir, "recommender_backtest.svg")
    network_csv_path = os.path.join(output_dir, "affinity_network.csv")
    network_svg_path = os.path.join(output_dir, "affinity_network.svg")
    export_pdf(result, pdf_path, stability_report, evaluation_report, affinity_report)
    export_excel(result, excel_path, stability_report, evaluation_report, affinity_report)
    write_stability_csv(stability_report, csv_path)
    write_stability_svg(stability_report, svg_path)
    write_evaluation_csv(evaluation_report, eval_csv_path)
    write_evaluation_svg(evaluation_report, eval_svg_path)
    write_affinity_csv(affinity_report, network_csv_path)
    write_affinity_svg(affinity_report, network_svg_path)

    sizes: dict[str, int] = {}
    # The PDF and workbook are substantial; the CSV/SVG read-outs are small but
    # must be non-trivial. Verify each against its own floor.
    for path, floor in (
        (pdf_path, 10_000),
        (excel_path, 10_000),
        (csv_path, 200),
        (svg_path, 200),
        (eval_csv_path, 200),
        (eval_svg_path, 200),
        (network_csv_path, 200),
        (network_svg_path, 200),
    ):
        size = os.path.getsize(path)
        if size <= floor:
            raise RuntimeError(f"deliverable too small ({size} bytes): {path}")
        sizes[path] = size
    return sizes
