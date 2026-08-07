from openpyxl import load_workbook

from basket.exports import write_deliverables


def test_deliverables_written_and_nonempty(tmp_path):
    sizes = write_deliverables(output_dir=str(tmp_path), n_baskets=1500, seed=42)
    paths = sorted(sizes)
    assert len(paths) == 6
    assert any(path.endswith(".pdf") for path in paths)
    assert any(path.endswith(".xlsx") for path in paths)
    assert any(path.endswith("rule_stability.csv") for path in paths)
    assert any(path.endswith("rule_stability.svg") for path in paths)
    assert any(path.endswith("recommender_backtest.csv") for path in paths)
    assert any(path.endswith("recommender_backtest.svg") for path in paths)
    for path, size in sizes.items():
        assert size > 200, f"{path} is only {size} bytes"

    excel_path = next(path for path in paths if path.endswith(".xlsx"))
    workbook = load_workbook(excel_path, read_only=True)
    assert workbook.sheetnames == [
        "Rules", "Itemsets", "Segments", "Recommendations", "Stability", "Evaluation"
    ]
    workbook.close()


def test_stability_outputs_deterministic(tmp_path):
    """CSV and SVG must be byte-identical across re-runs (fixed seed, no clock)."""
    first = tmp_path / "a"
    second = tmp_path / "b"
    write_deliverables(output_dir=str(first), n_baskets=1500, seed=42)
    write_deliverables(output_dir=str(second), n_baskets=1500, seed=42)
    for name in (
        "rule_stability.csv",
        "rule_stability.svg",
        "recommender_backtest.csv",
        "recommender_backtest.svg",
    ):
        assert (first / name).read_bytes() == (second / name).read_bytes(), name
